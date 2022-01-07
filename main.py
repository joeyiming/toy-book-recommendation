from os import read
from types import resolve_bases
from openpyxl import Workbook, load_workbook
import random
import pyperclip
import sys


# 一个过分简单的测试函数
def test(workBook: Workbook)->bool:
    sheet = workBook.active
    a1 = sheet['A1']
    return bool(a1)

def fetchData(storeData: list, workBook: Workbook,filterAttr='筛选标记')->None:
    """获取数据，如果筛选标记不为 0 则存储。

    Args:
        storeData (list): 用于存储数据的列表
        workBook (Workbook): 用于获取数据的WorkBook
    """

    sheet = workBook.active
    # 获取列首属性名
    attrList = []
    for cell in sheet[1]:
        attrList.append(cell.value)
    for row in sheet.iter_rows(min_row=2):
        newBook = {}
        for index, cell in enumerate(row):
            newBook[attrList[index]] = cell.value
        # 筛选操作
        if newBook[filterAttr]!=0:
            storeData.append(newBook)
    # print('总条数:',len(storeData))

def copyLink(book:dict,linkAttrName='相关链接')->None:
    if book[linkAttrName]:
        pyperclip.copy(book[linkAttrName])
    print(f'已将{linkAttrName}复制到粘贴板\n')

def getLongInfo(book: dict,displayWidth=40,notDisplayAttrNames=['筛选标记','图书名','主观评级','主题']) -> str:
    """获取书本的格式化介绍性信息

    Args:
        book (dict): 以字典形式存储的书本
        notDisPlayAttrNames: 不予显示的属性名列表。

    Returns:
        str: 返回格式化的介绍信息
    """
    nameLine = getBookName(book).center(displayWidth)+'\n'
    spaceLine = ' '*displayWidth+'\n'
    dashLine = '-'*displayWidth*2+'\n'
    aftername = ''
    for key,value in book.items():
        if str(key) in notDisplayAttrNames:
            continue
        else:
            if value:
                # keyline = str(key).center(displayWidth)+'\n'
                valueline = str(value).center(displayWidth)+'\n'
                aftername+=spaceLine+valueline
    result = dashLine+spaceLine+nameLine+aftername+spaceLine+dashLine
    return result

def getBookName(book:dict,nameAttr='图书名')->str:
    return '《'+book[nameAttr]+'》'


def randomMode(data: list,filename:str) -> None:
    """全随机模式：从整个图书列表中随机抽取一本书，打印到控制台并保存为文本文件。

    Args:
        data (list): 图书列表
        filename (str): 输出结果文件文件名
    """
    choosenBook = random.choice(data)
    result = getLongInfo(choosenBook)
    print(result)
    copyLink(choosenBook)
    with open(filename,'w+',encoding='utf8') as f:
        f.write(result)

def filterMode(data:list,filename:str,condition=1):
    """筛选模式：根据特定条件对图书列表筛选，并从筛选结果中随机抽取一本书，打印到控制台并保存为文本文件。

    Args:
        data (list): 图书列表
        filename (str): 输出结果文件文件名
        condition (int): 条件代码
            1: 主题限定为技术
            2: 主题不含技术
    """
    filteredData = []

    # condition: 1
    # 技术书 
    if condition==1:
        for book in data:
            if book['主题'].count('技术')>0:
                filteredData.append(book)
    
    # condition: 2
    # 非技术书 
    if condition==2:
        for book in data:
            if book['主题'].count('技术')==0:
                filteredData.append(book)

    # print('筛选后条数',len(filteredData))
    result=''
    if len(filteredData)>0:
        choosenBook = random.choice(filteredData)
        result = getLongInfo(choosenBook)
        print(result)
    else:
        print('筛选结果为空。')
    with open(filename,'w+',encoding='utf8') as f:
        f.write(result)



def main():
    OPENFILENAME = 'book-list.xlsx'
    SAVAFILENAME = '抽奖结果.txt'
    readBook = load_workbook(OPENFILENAME)
    bookList = []
    if not test(readBook):
        # 过分简单的异常处理
        print('运行错误')
        return
    fetchData(bookList, readBook)

    # 如果运行命令不带额外参数，进入互动，否则跳过互动
    if (len(sys.argv)<2):
        print('你好，请问有筛选条件么？任何书还是就技术书呢？\n')
        answer=input('0: 任何书 1: 就技术书 2:不要技术书 （直接回车，默认为0）：').strip()
        if answer=='':
            answer=0
        else:
            answer=int(answer)
    else:
        answer=int(sys.argv[1])
    if answer==0:
        randomMode(bookList,SAVAFILENAME)
    else:
        filterMode(bookList,SAVAFILENAME,answer)
    print('很高兴能帮到你，再见。\n')



if __name__ == '__main__':
    main()
