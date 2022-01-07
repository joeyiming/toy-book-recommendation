from os import read
from types import resolve_bases
from openpyxl import Workbook, load_workbook
import random


# 一个过分简单的测试函数
def test(workBook: Workbook)->bool:
    sheet = workBook.active
    a1 = sheet['A1']
    return bool(a1)

def fetchData(storeData: list, workBook: Workbook)->None:
    """获取数据，并存储在函数外的列表中

    Args:
        storeData (list): 用于存储数据的列表
        workBook (Workbook): 用于获取数据的WorkBook
    """

    sheet = workBook.active
    # 获取列首属性名
    attrList = []
    for cell in sheet[1]:
        attrList.append(cell.value)
    # attrList=['Flag','Name','Stars','Themes','Memo','Link'] 
    # print(attrList)
    for row in sheet.iter_rows(min_row=2):
        newBook = {}
        for index, cell in enumerate(row):
            newBook[attrList[index]] = cell.value
        storeData.append(newBook)

def getLongInfo(book: dict,displayWidth=40,notDisplayAttrNames=['筛选标记','图书名','主观评级','主题','相关链接']) -> str:
    """获取书本的格式化介绍性信息

    Args:
        book (dict): 以字典形式存储的书本
        notDisPlayAttrNames: 不予显示的属性名列表。

    Returns:
        str: 返回格式化的介绍信息
    """
    nameLine = getBookName(book).center(displayWidth)+'\n'
    spaceLine = ' '*displayWidth+'\n'
    dashLine = '-'*displayWidth*2+'\n'
    aftername = ''
    for key,value in book.items():
        if str(key) in notDisplayAttrNames:
            continue
        else:
            if value:
                # keyline = str(key).center(displayWidth)+'\n'
                valueline = str(value).center(displayWidth)+'\n'
                aftername+=spaceLine+valueline
    result = dashLine+spaceLine+nameLine+aftername+spaceLine+dashLine
    return result

def getBookName(book:dict,nameAttr='图书名')->str:
    return '《'+book[nameAttr]+'》'


def randomMode(data: list,filename:str) -> None:
    """全随机模式：从整个图书列表中随机抽取一本书，打印到控制台并保存为文本文件。

    Args:
        data (list): 图书列表
    """
    choosenBook = random.choice(data)
    result = getLongInfo(choosenBook)
    print(result)
    with open(filename,'w+',encoding='utf8') as f:
        f.write(result)



def main() -> None:
    OPENFILENAME = 'book-list.xlsx'
    SAVAFILENAME = '抽奖结果.txt'
    readBook = load_workbook(OPENFILENAME)
    bookList = []
    if not test(readBook):
        # 过分简单的异常处理
        print('运行错误')
        return

    fetchData(bookList, readBook)
    randomMode(bookList,SAVAFILENAME)


if __name__ == '__main__':
    main()
